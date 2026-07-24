import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_FILE = PROJECT_ROOT / "data" / "spieler.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "data" / "spieler.json"

EXPECTED_COLUMNS = [
    "id",
    "name",
    "nummer",
    "position",
    "seit",
    "min_liga",
    "tore_liga",
    "vorlagen_liga",
    "gelb_liga",
    "gelbrot_liga",
    "rot_liga",
    "min_zwoote",
    "tore_zwoote",
    "vorlagen_zwoote",
    "gelb_zwoote",
    "gelbrot_zwoote",
    "rot_zwoote",
]

NUMERIC_FIELDS = {
    "nummer",
    "min_liga",
    "tore_liga",
    "vorlagen_liga",
    "gelb_liga",
    "gelbrot_liga",
    "rot_liga",
    "min_zwoote",
    "tore_zwoote",
    "vorlagen_zwoote",
    "gelb_zwoote",
    "gelbrot_zwoote",
    "rot_zwoote",
}


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip()


def convert_value(key, value):
    if value is None:
        return 0 if key in NUMERIC_FIELDS else ""

    if key in NUMERIC_FIELDS:
        if value == "":
            return 0
        try:
            return int(value)
        except (ValueError, TypeError):
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return 0

    return str(value).strip()


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Excel-Datei nicht gefunden: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Das Excel-Blatt ist leer.")

    headers = [normalize_header(cell) for cell in rows[0]]

    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Fehlende Spalten: {', '.join(missing)}")

    header_index = {header: idx for idx, header in enumerate(headers)}

    players = []
    for row in rows[1:]:
        player_id = row[header_index["id"]] if header_index["id"] < len(row) else None
        if player_id is None or str(player_id).strip() == "":
            continue

        player = {}
        for key in EXPECTED_COLUMNS:
            idx = header_index[key]
            raw_value = row[idx] if idx < len(row) else None
            player[key] = convert_value(key, raw_value)

        players.append(player)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_FILE.open("w", encoding="utf-8") as f:
        json.dump(players, f, ensure_ascii=False, indent=2)

    print(f"{len(players)} Spieler nach {OUTPUT_FILE} exportiert.")


if __name__ == "__main__":
    main()